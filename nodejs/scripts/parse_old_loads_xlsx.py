import json
import sys
from openpyxl import load_workbook


def to_text(value):
    if value is None:
        return ""
    return str(value).strip()


def main():
    if len(sys.argv) < 2:
        raise RuntimeError("Expected workbook path argument")

    workbook_path = sys.argv[1]
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [to_text(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)]
    header_map = {h: idx + 1 for idx, h in enumerate(headers) if h}

    rows = []
    for r in range(2, ws.max_row + 1):
        record = {}
        empty = True
        for header, col in header_map.items():
            value = ws.cell(row=r, column=col).value
            if value is not None and to_text(value) != "":
                empty = False
            record[header] = to_text(value)
        if not empty:
            rows.append(record)

    print(json.dumps({
        "sheetName": ws.title,
        "headers": headers,
        "rows": rows
    }))


if __name__ == "__main__":
    main()
